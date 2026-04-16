"""
재고 대시보드 V4 (Flask, 진입 파일명 dashboard_v2.py 유지)
포트(로컬): 5003
"""
from __future__ import annotations

import datetime as dt
import os
import re
import traceback
from urllib.parse import urlencode
from collections import defaultdict
from functools import wraps
from io import BytesIO, StringIO
from typing import Any, Callable, Optional

import pandas as pd
import plotly.express as px
import requests
from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
    send_file,
)
from flask_caching import Cache

from inventory_core import (
    avg_daily_usage_from_history,
    compute_daily_change,
    get_conn,
    init_db,
    load_history,
    load_latest,
    normalize_excel,
    reorder_suggestion,
    upsert_snapshot,
    update_channel_stock,
    update_warehouse_stock,
    update_distribution_note,
)

# [최적화] 뷰 단에서 Cache 인스턴스에 데코레이터 연결
cache = Cache()


def _invalidate_snapshot_caches() -> None:
    """스냅샷 데이터 변경 후 dashboard_base·item_summary_* 만 무효화 (전체 clear 지양)."""
    cache.delete("dashboard_base")
    _backend = getattr(cache.cache, "_cache", None)
    if _backend is None:
        cache.clear()
        # TODO: item_summary 키만 삭제하도록 개선 필요
        return
    _keys = [k for k in _backend.keys() if str(k).startswith("item_summary_")]
    if _keys:
        cache.delete_many(*_keys)


APP_TITLE = "재고 대시보드 V4"
DEFAULT_PASSWORD = "1234"
_pw_cache = {"value": None}

# DB에 gsheet_url이 없을 때 업로드 폼·동기화에 쓰는 기본 스프레드시트
DEFAULT_GSHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1lphWwIG146kEZBxVP5_STqFDH7yOxeZHI4tXpedgT3g/edit?usp=sharing"
)
DEFAULT_GSHEET_SHEET = "재고판매현황"


# ---------------------------------------------------------------------------
# 구글 시트 헬퍼
# ---------------------------------------------------------------------------

def _extract_spreadsheet_id(url_or_id: str) -> str:
    """URL 또는 ID 문자열에서 구글 스프레드시트 ID 추출"""
    import re
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_or_id)
    return m.group(1) if m else url_or_id.strip()


def fetch_gsheet_as_dataframe(url_or_id: str, sheet_name: str = "재고판매현황") -> pd.DataFrame:
    # [최적화] 공개 CSV는 requests 타임아웃 + StringIO 파싱
    """구글 시트 → DataFrame

    1차: 공개 시트 CSV export URL로 requests(타임아웃) + pd.read_csv(StringIO)
    2차(실패 시): 환경변수 GSHEET_SERVICE_ACCOUNT_PATH 또는
                  GSHEET_SERVICE_ACCOUNT_JSON 으로 gspread 서비스 계정 인증
    """
    import urllib.parse

    spreadsheet_id = _extract_spreadsheet_id(url_or_id)
    encoded_sheet = urllib.parse.quote(sheet_name)

    # 공개 시트 CSV 방식
    csv_url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        f"/gviz/tq?tqx=out:csv&sheet={encoded_sheet}"
    )
    try:
        resp = requests.get(csv_url, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(StringIO(resp.text))
        if df.empty:
            raise ValueError("시트가 비어 있습니다.")
        return df
    except Exception:
        pass  # 비공개 시트면 아래 서비스 계정 방식으로 재시도

    # 서비스 계정 방식 (환경변수 설정 시)
    sa_path = os.environ.get("GSHEET_SERVICE_ACCOUNT_PATH", "").strip()
    sa_json_str = os.environ.get("GSHEET_SERVICE_ACCOUNT_JSON", "").strip()

    if not sa_path and not sa_json_str:
        raise ValueError(
            "구글 시트 읽기에 실패했습니다. "
            "시트가 '링크가 있는 누구나 볼 수 있음'으로 공유되어 있는지 확인하거나, "
            "비공개 시트라면 환경변수 GSHEET_SERVICE_ACCOUNT_PATH 또는 "
            "GSHEET_SERVICE_ACCOUNT_JSON 을 설정하세요."
        )

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ValueError(
            "비공개 시트 접근을 위해 'pip install gspread google-auth' 가 필요합니다."
        )

    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    try:
        if sa_json_str:
            import json
            creds = Credentials.from_service_account_info(
                json.loads(sa_json_str), scopes=scopes
            )
        else:
            creds = Credentials.from_service_account_file(sa_path, scopes=scopes)

        gc = gspread.authorize(creds)
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        data = ws.get_all_records()
        if not data:
            raise ValueError("시트가 비어 있습니다.")
        return pd.DataFrame(data)
    except Exception as sa_err:
        raise ValueError(f"서비스 계정으로 구글 시트 읽기 실패: {sa_err}")

# 팀 배포 모드: True면 초기화 기능 비활성화, /test 비노출, 500 에러 시 상세 미표시
DEPLOY_MODE = os.environ.get("DEPLOY_MODE", "").strip().lower() in ("1", "true", "yes")


def create_app() -> Flask:
    # [최적화] SimpleCache + DB 1회 초기화(init_db)
    """Flask 앱 생성 및 설정"""
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key-change-me-v2")
    app.config["PROPAGATE_EXCEPTIONS"] = True
    app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB 제한
    app.config["CACHE_TYPE"] = "SimpleCache"
    app.config["CACHE_DEFAULT_TIMEOUT"] = 300
    cache.init_app(app)
    init_db()
    return app


app = create_app()

# 표시용 상품명: 마지막 `_` 뒤에 한글이 없으면(영문·숫자·기호 접미사로 간주) 해당 `_…` 구간 제거
_HANGUL_IN_PRODUCT = re.compile(r"[가-힣ㄱ-ㅎㅏ-ㅣ]")


def display_product_name_ui(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    idx = s.rfind("_")
    if idx > 0:
        tail = s[idx + 1 :]
        if tail and not _HANGUL_IN_PRODUCT.search(tail):
            return s[:idx].rstrip()
    return s


app.jinja_env.filters["product_name_ui"] = display_product_name_ui


@app.errorhandler(500)
def internal_error(e):
    """500 에러 핸들러 (배포 모드에서는 상세 미표시)"""
    if DEPLOY_MODE:
        return (
            "<h1>500 Internal Server Error</h1>"
            "<p>일시적인 오류가 발생했습니다. 잠시 후 다시 시도해 주세요.</p>",
            500,
        )
    import traceback
    tb = traceback.format_exc()
    return (
        f"<h1>500 Internal Server Error</h1>"
        f"<pre style='background:#fee;padding:1em;border-radius:8px;'>{tb}</pre>",
        500,
    )


@app.errorhandler(404)
def not_found(e):
    """404 에러 핸들러"""
    return "<h1>404 Not Found</h1><p>요청한 페이지를 찾을 수 없습니다.</p>", 404


@app.context_processor
def inject_deploy_config():
    """템플릿에 배포 설정 전달 (초기화 버튼 노출 여부)"""
    return {"show_clear_data": not DEPLOY_MODE}


@app.context_processor
def inject_gsheet_sync_nav():
    """네비 DB 최신화 버튼용: 오늘 날짜·동기화 가능 URL(기본값 포함) 여부"""
    st = _get_gsheet_settings()
    return {
        "gsheet_sync_snapshot_default": dt.date.today().isoformat(),
        "has_gsheet_saved": bool((st.get("url") or "").strip()),
    }


@app.route("/test")
def test():
    """서버 상태 확인 (배포 모드에서는 비노출)"""
    if DEPLOY_MODE:
        abort(404)
    return "<h1>OK</h1><p>✅ 대시보드 V4 서버 정상 작동중 (포트: 5003)</p>"


def _get_gsheet_settings() -> dict:
    """구글 시트 설정. DB에 URL이 없으면 DEFAULT_GSHEET_* 로 채움. saved=DB에 URL이 있는지."""
    url_db, sheet_db = "", ""
    try:
        conn = get_conn()
        cur = conn.execute("SELECT key, value FROM settings WHERE key IN ('gsheet_url', 'gsheet_sheet')")
        rows = {row[0]: row[1] for row in cur.fetchall()}
        url_db = (rows.get("gsheet_url") or "").strip()
        sheet_db = (rows.get("gsheet_sheet") or "").strip()
    except Exception:
        pass
    saved = bool(url_db)
    sheet = sheet_db or DEFAULT_GSHEET_SHEET
    url = url_db or DEFAULT_GSHEET_URL
    return {"url": url, "sheet": sheet, "saved": saved}


def _set_gsheet_settings(url: str, sheet: str) -> None:
    """DB에 구글 시트 연동 설정 저장"""
    conn = get_conn()
    conn.executemany(
        "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
        [("gsheet_url", url.strip()), ("gsheet_sheet", (sheet or "재고판매현황").strip())],
    )
    conn.commit()


def _get_password_from_db() -> str:
    """DB에서 비밀번호 조회 (없으면 기본값으로 초기화)"""
    try:
        conn = get_conn()
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        cur = conn.execute("SELECT value FROM settings WHERE key = 'password'")
        row = cur.fetchone()
        
        if row and row[0]:
            return str(row[0])
        
        # 기본 비밀번호로 초기화
        conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES ('password', ?)",
            (DEFAULT_PASSWORD,),
        )
        conn.commit()
        return DEFAULT_PASSWORD
    except Exception as e:
        print(f"[ERROR] 비밀번호 조회 실패: {e}")
        return DEFAULT_PASSWORD


def _set_password_in_db(new_password: str) -> None:
    """DB에 비밀번호 저장"""
    new_password = (new_password or "").strip()
    if not new_password:
        raise ValueError("비밀번호는 빈 값일 수 없습니다.")
    
    try:
        conn = get_conn()
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES ('password', ?)",
            (new_password,),
        )
        conn.commit()
        _pw_cache["value"] = new_password
    except Exception as e:
        print(f"[ERROR] 비밀번호 저장 실패: {e}")
        raise


def _expected_password() -> str:
    """로그인에 사용할 현재 비밀번호 반환"""
    if _pw_cache["value"] is not None:
        return _pw_cache["value"]
    _pw_cache["value"] = _get_password_from_db()
    return _pw_cache["value"]


def login_required(view: Callable[..., Any]) -> Callable[..., Any]:
    """로그인 필수 데코레이터"""
    @wraps(view)
    def wrapper(*args: Any, **kwargs: Any) -> Any:
        expected = _expected_password()
        if expected and not session.get("authed"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapper


@app.get("/login")
def login() -> str:
    """로그인 페이지"""
    if session.get("authed"):
        return redirect(url_for("dashboard"))
    return render_template("login.html", title=APP_TITLE)


@app.post("/login")
def login_post():
    """로그인 처리"""
    expected = _expected_password()
    if not expected:
        session["authed"] = True
        return redirect(url_for("dashboard"))
    
    pw = (request.form.get("password") or "").strip()
    if pw and pw == expected:
        session["authed"] = True
        return redirect(request.args.get("next") or url_for("dashboard"))
    
    flash("비밀번호가 올바르지 않습니다.", "danger")
    return redirect(url_for("login"))


@app.get("/logout")
def logout():
    """로그아웃"""
    session.clear()
    return redirect(url_for("login"))


@app.get("/backup")
@login_required
def backup_page():
    """백업 페이지"""
    return render_template("backup.html", title=APP_TITLE)


def _build_fillup_excel(df: pd.DataFrame, latest_date: str) -> BytesIO:
    """필업지 엑셀(아소트 BOX + 솔리드 PCS) BytesIO 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"
    HM_HDR = "2E75B6"
    BP_HDR = "375623"
    HM_DATA = "DEEAF1"
    BP_DATA = "E2EFDA"
    HM_ALT = "EBF3FA"
    BP_ALT = "EAF4E7"
    ID_HDR = "BDD7EE"
    ID_FC = "1F3864"
    AN_HDR = "595959"
    INPUT = "FFFDE7"
    ROW_ALT = "F5F5F5"

    STATUS_ORDER = {
        "긴급필업": 0,
        "재고없음": 1,
        "필업필요": 2,
        "체크필요": 3,
        "저재고": 4,
        "필업검토": 5,
        "정상": 6,
    }
    STATUS_STYLE = {
        "긴급필업": ("FF0000", "FFE6E6"),
        "재고없음": ("FFFFFF", "333333"),
        "필업필요": ("C55A11", "FDEBD7"),
        "체크필요": ("0070C0", "DEEBF7"),
        "저재고": ("7F6000", "FFF2CC"),
        "필업검토": ("595959", "F2F2F2"),
        "정상": ("375623", "E2EFDA"),
    }

    thin = Side(style="thin", color="BBBBBB")
    thick = Side(style="medium", color="555555")
    NB = Border(left=thin, right=thin, top=thin, bottom=thin)
    IB = Border(left=thick, right=thick, top=thick, bottom=thick)

    def _c(ws, row, col, val="", bg="FFFFFF", fc="000000", bold=False, sz=9, align="center", wrap=False, border=NB):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, size=sz, name="Arial", color=fc)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border = border
        return cell

    wb = Workbook()

    # ── 시트1: 아소트 필업지 ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "아소트 필업지 (BOX단위)"

    df_a = df[(df["warehouse1_assort"] > 0) | (df["warehouse2_assort"] > 0)].copy()

    assort_rows = []
    for (style, color), grp in df_a.groupby(["style_code", "color_code"], sort=False):
        pcs_per_box = int(grp["assort_ratio"].sum())
        hm_assort_pcs = int(grp["warehouse1_assort"].sum())
        bp_assort_pcs = int(grp["warehouse2_assort"].sum())
        total_pcs = hm_assort_pcs + bp_assort_pcs

        if pcs_per_box > 0:
            hm_box = round(hm_assort_pcs / pcs_per_box, 1)
            bp_box = round(bp_assort_pcs / pcs_per_box, 1)
            unit = "BOX"
        else:
            hm_box = hm_assort_pcs
            bp_box = bp_assort_pcs
            unit = "PCS"

        stock_sum = int(grp["stock"].sum())
        online_sum = int(grp["online_stock"].sum())
        channel_sum = int(grp["channel_stock"].sum())
        daily_sum = round(float(grp["daily_sales_7d"].sum()), 1)
        days_out = round((stock_sum + total_pcs) / daily_sum, 1) if daily_sum > 0 else 999.0
        weeks_out = round(days_out / 7, 1) if days_out < 999 else "-"
        suggested = int(grp["suggested_order_qty"].sum())
        season = grp["season"].iloc[0]
        name = str(grp["name"].iloc[0]) if "name" in grp.columns else ""

        worst_status = min(grp["status"].tolist(), key=lambda s: STATUS_ORDER.get(s, 9))

        assort_rows.append(
            {
                "style": style,
                "color": color,
                "name": name,
                "daily": daily_sum,
                "stock": stock_sum,
                "online_stock": online_sum,
                "channel_stock": channel_sum,
                "season": season,
                "status": worst_status,
                "hm_box": hm_box,
                "bp_box": bp_box,
                "unit": unit,
                "pcs_per_box": pcs_per_box if pcs_per_box > 0 else "-",
                "total_pcs": total_pcs,
                "days_out": days_out,
                "weeks_out": weeks_out,
                "suggested": suggested,
            }
        )

    assort_rows.sort(key=lambda x: x["daily"], reverse=True)

    ws1.merge_cells("A1:R1")
    _c(
        ws1,
        1,
        1,
        f"📋 아소트 필업지  |  스타일+컬러 / BOX단위  |  기준일: {latest_date}",
        NAVY,
        "FFFFFF",
        bold=True,
        sz=12,
        align="left",
    )

    sections = [
        ("상품 정보", 1, 9, ID_HDR, ID_FC),
        ("📦 항만 물류센터", 10, 12, HM_HDR, "FFFFFF"),
        ("📦 부평 물류센터", 13, 15, BP_HDR, "FFFFFF"),
        ("재고 분석", 16, 18, AN_HDR, "FFFFFF"),
    ]
    for label, s, e, bg, fc in sections:
        ws1.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
        _c(ws1, 2, s, label, bg, fc, bold=True, sz=10, border=IB)
        for ci in range(s + 1, e + 1):
            ws1.cell(row=2, column=ci).fill = PatternFill("solid", fgColor=bg)
            ws1.cell(row=2, column=ci).border = IB

    hdrs1 = [
        (1, "스타일코드", ID_HDR, ID_FC),
        (2, "컬러코드", ID_HDR, ID_FC),
        (3, "상품명", ID_HDR, ID_FC),
        (4, "시즌", ID_HDR, ID_FC),
        (5, "상태", ID_HDR, ID_FC),
        (6, "일평균\n판매", ID_HDR, ID_FC),
        (7, "현재재고", ID_HDR, ID_FC),
        (8, "공홈재고\n(현재-매장)", ID_HDR, "0070C0"),
        (9, "매장재고", ID_HDR, ID_FC),
        (10, "가용\nBOX수", HM_HDR, "FFFFFF"),
        (11, "PCS/BOX", HM_HDR, "FFFFFF"),
        (12, "분배요청\n[BOX]", HM_HDR, "FFFFFF"),
        (13, "가용\nBOX수", BP_HDR, "FFFFFF"),
        (14, "PCS/BOX", BP_HDR, "FFFFFF"),
        (15, "분배요청\n[BOX]", BP_HDR, "FFFFFF"),
        (16, "아소트\n총PCS", AN_HDR, "FFFFFF"),
        (17, "소진예상일", AN_HDR, "FFFFFF"),
        (18, "필업제안\n수량", AN_HDR, "FFFFFF"),
    ]
    for ci, label, bg, fc in hdrs1:
        _c(ws1, 3, ci, label, bg, fc, bold=True, sz=9, wrap=True)

    for ri, r in enumerate(assort_rows, start=4):
        even = ri % 2 == 0
        bg_id = ROW_ALT if even else "FFFFFF"
        bg_hm = HM_ALT if even else HM_DATA
        bg_bp = BP_ALT if even else BP_DATA
        sc = STATUS_STYLE.get(r["status"], ("000000", "FFFFFF"))

        _c(ws1, ri, 1, r["style"], bg_id, align="left")
        _c(ws1, ri, 2, r["color"], bg_id)
        _c(ws1, ri, 3, r["name"], bg_id, align="left", wrap=True)
        _c(ws1, ri, 4, r["season"], bg_id)

        _c(ws1, ri, 5, r["status"], sc[1], sc[0], bold=True)

        _c(ws1, ri, 6, r["daily"], bg_id, "0070C0", bold=True)
        _c(ws1, ri, 7, r["stock"], bg_id, fc="FF0000" if r["stock"] == 0 else "000000", bold=(r["stock"] == 0))

        _c(
            ws1,
            ri,
            8,
            r["online_stock"],
            bg_id,
            fc="FF0000" if r["online_stock"] == 0 else "0070C0",
            bold=True,
            border=IB,
        )

        _c(ws1, ri, 9, r["channel_stock"], bg_id, "666666")

        _c(ws1, ri, 10, r["hm_box"], bg_hm, bold=True)
        _c(ws1, ri, 11, r["pcs_per_box"], bg_hm, "555555")
        _c(ws1, ri, 12, "", INPUT, border=IB)

        _c(ws1, ri, 13, r["bp_box"], bg_bp, bold=True)
        _c(ws1, ri, 14, r["pcs_per_box"], bg_bp, "555555")
        _c(ws1, ri, 15, "", INPUT, border=IB)

        _c(ws1, ri, 16, r["total_pcs"], bg_id)

        d = r["days_out"]
        d_fc = "FF0000" if d < 7 else ("C55A11" if d < 14 else "000000")
        _c(ws1, ri, 17, "-" if d >= 999 else d, bg_id, d_fc, bold=(d < 14))

        sq = r["suggested"]
        _c(ws1, ri, 18, sq, "FFD7D7" if sq > 0 else bg_id, "C00000" if sq > 0 else "000000", bold=(sq > 0))

        ws1.row_dimensions[ri].height = 16

    col_w1 = [14, 7, 36, 6, 9, 9, 8, 12, 8, 10, 9, 13, 10, 9, 13, 10, 10, 11]
    for i, w in enumerate(col_w1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 34
    ws1.freeze_panes = "A4"

    # ── 시트2: 솔리드 필업지 ─────────────────────────────────────
    ws2 = wb.create_sheet("솔리드 필업지 (PCS단위)")

    df_s = df[(df["warehouse1_solid"] > 0) | (df["warehouse2_solid"] > 0)].copy()

    ws2.merge_cells("A1:P1")
    _c(
        ws2,
        1,
        1,
        f"📋 솔리드 필업지  |  SKU / PCS단위  |  기준일: {latest_date}",
        NAVY,
        "FFFFFF",
        bold=True,
        sz=12,
        align="left",
    )

    sections2 = [
        ("상품 정보", 1, 9, ID_HDR, ID_FC),
        ("📦 항만 물류센터", 10, 11, HM_HDR, "FFFFFF"),
        ("📦 부평 물류센터", 12, 13, BP_HDR, "FFFFFF"),
        ("재고 분석", 14, 16, AN_HDR, "FFFFFF"),
    ]
    for label, s, e, bg, fc in sections2:
        ws2.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
        _c(ws2, 2, s, label, bg, fc, bold=True, sz=10, border=IB)
        for ci in range(s + 1, e + 1):
            ws2.cell(row=2, column=ci).fill = PatternFill("solid", fgColor=bg)
            ws2.cell(row=2, column=ci).border = IB

    hdrs2 = [
        (1, "SKU", ID_HDR, ID_FC),
        (2, "상품명", ID_HDR, ID_FC),
        (3, "사이즈", ID_HDR, ID_FC),
        (4, "시즌", ID_HDR, ID_FC),
        (5, "상태", ID_HDR, ID_FC),
        (6, "일평균\n판매", ID_HDR, ID_FC),
        (7, "현재재고", ID_HDR, ID_FC),
        (8, "공홈재고\n(현재-매장)", ID_HDR, "0070C0"),
        (9, "매장재고", ID_HDR, ID_FC),
        (10, "항만\n가용PCS", HM_HDR, "FFFFFF"),
        (11, "분배요청\n[PCS]", HM_HDR, "FFFFFF"),
        (12, "부평\n가용PCS", BP_HDR, "FFFFFF"),
        (13, "분배요청\n[PCS]", BP_HDR, "FFFFFF"),
        (14, "물류재고합", AN_HDR, "FFFFFF"),
        (15, "소진예상일", AN_HDR, "FFFFFF"),
        (16, "필업제안\n수량", AN_HDR, "FFFFFF"),
    ]
    for ci, label, bg, fc in hdrs2:
        _c(ws2, 3, ci, label, bg, fc, bold=True, sz=9, wrap=True)

    for ri, (_, srow) in enumerate(df_s.iterrows(), start=4):
        even = ri % 2 == 0
        bg_id = ROW_ALT if even else "FFFFFF"
        bg_hm = HM_ALT if even else HM_DATA
        bg_bp = BP_ALT if even else BP_DATA

        sku = str(srow.get("sku", "") or "")
        name = str(srow.get("name", "") or "")
        size = str(srow.get("size_code", "") or "")
        season = str(srow.get("season", "") or "")
        status = str(srow.get("status", "정상") or "정상")
        daily = float(srow.get("daily_sales_7d") or 0)
        stock = int(srow.get("stock") or 0)
        online = int(srow.get("online_stock") or 0)
        channel = int(srow.get("channel_stock") or 0)
        hm_s = int(srow.get("warehouse1_solid") or 0)
        bp_s = int(srow.get("warehouse2_solid") or 0)
        wh_tot = hm_s + bp_s
        d = float(srow.get("days_until_out") or 999)
        sq = int(srow.get("suggested_order_qty") or 0)

        sc = STATUS_STYLE.get(status, ("000000", "FFFFFF"))

        _c(ws2, ri, 1, sku, bg_id, align="left", sz=8)
        _c(ws2, ri, 2, name, bg_id, align="left", wrap=True)
        _c(ws2, ri, 3, size, bg_id)
        _c(ws2, ri, 4, season, bg_id)
        _c(ws2, ri, 5, status, sc[1], sc[0], bold=True)
        _c(ws2, ri, 6, round(daily, 1), bg_id, "0070C0", bold=True)
        _c(ws2, ri, 7, stock, bg_id, fc="FF0000" if stock == 0 else "000000", bold=(stock == 0))

        _c(
            ws2,
            ri,
            8,
            online,
            bg_id,
            fc="FF0000" if online == 0 else "0070C0",
            bold=True,
            border=IB,
        )

        _c(ws2, ri, 9, channel, bg_id, "666666")
        _c(ws2, ri, 10, hm_s, bg_hm, bold=True)
        _c(ws2, ri, 11, "", INPUT, border=IB)
        _c(ws2, ri, 12, bp_s, bg_bp, bold=True)
        _c(ws2, ri, 13, "", INPUT, border=IB)
        _c(ws2, ri, 14, wh_tot, bg_id)

        d_fc = "FF0000" if d < 7 else ("C55A11" if d < 14 else "000000")
        _c(ws2, ri, 15, "-" if d >= 999 else d, bg_id, d_fc, bold=(d < 14))
        _c(ws2, ri, 16, sq, "FFD7D7" if sq > 0 else bg_id, "C00000" if sq > 0 else "000000", bold=(sq > 0))

        ws2.row_dimensions[ri].height = 16

    col_w2 = [18, 36, 7, 6, 9, 9, 8, 12, 8, 10, 13, 10, 13, 10, 10, 11]
    for i, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.row_dimensions[1].height = 24
    ws2.row_dimensions[2].height = 20
    ws2.row_dimensions[3].height = 34
    ws2.freeze_panes = "A4"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/export/fillup")
@login_required
def export_fillup():
    conn = get_conn()
    try:
        latest_date, latest = load_latest(conn)
        if latest_date is None or latest.empty:
            flash("출력할 데이터가 없습니다.", "warning")
            return redirect(url_for("dashboard"))

        wh = pd.read_sql_query(
            """
            SELECT sku,
                   COALESCE(warehouse1_solid, 0) AS warehouse1_solid,
                   COALESCE(warehouse1_assort, 0) AS warehouse1_assort,
                   COALESCE(warehouse2_solid, 0) AS warehouse2_solid,
                   COALESCE(warehouse2_assort, 0) AS warehouse2_assort
            FROM snapshots WHERE snapshot_date = ?
            """,
            conn,
            params=(latest_date,),
        )
        for c in ("warehouse1_solid", "warehouse1_assort", "warehouse2_solid", "warehouse2_assort"):
            if c in latest.columns:
                latest = latest.drop(columns=[c])
        latest = latest.merge(wh, on="sku", how="left")

        for col in (
            "sales_qty",
            "stock",
            "channel_stock",
            "warehouse_stock",
            "warehouse1_stock",
            "warehouse2_stock",
            "warehouse1_solid",
            "warehouse1_assort",
            "warehouse2_solid",
            "warehouse2_assort",
            "assort_ratio",
            "assort_box_count",
        ):
            if col not in latest.columns:
                latest[col] = 0
            latest[col] = pd.to_numeric(latest[col], errors="coerce").fillna(0).astype(int)

        latest["daily_sales_7d"] = (latest["sales_qty"] / 7.0).round(2)
        latest["total_available"] = latest["stock"] + latest["warehouse_stock"]
        latest["days_until_out"] = (
            latest["total_available"] / latest["daily_sales_7d"].replace(0, float("nan"))
        ).round(1).fillna(999)
        latest["suggested_order_qty"] = (
            (latest["daily_sales_7d"] * 14) - latest["total_available"]
        ).clip(lower=0).astype(int)

        latest["online_stock"] = (latest["stock"] - latest["channel_stock"]).clip(lower=0).astype(int)

        latest["style_code"] = latest["sku"].astype(str).str[:10]
        latest["color_code"] = latest["sku"].astype(str).str[10:12]
        latest["size_code"] = latest["sku"].astype(str).str[12:]
        latest["season"] = latest["sku"].astype(str).str[4:6]

        if "status" not in latest.columns:
            import numpy as np

            cond = [
                (latest["stock"] == 0) & (latest["daily_sales_7d"] > 0),
                (latest["stock"] == 0),
                (latest["daily_sales_7d"] > 0) & (latest["days_until_out"] < 7),
                (latest["stock"] <= 10) & (latest["daily_sales_7d"] > 0),
                (latest["days_until_out"] < 14) & (latest["daily_sales_7d"] > 0),
            ]
            choices = ["긴급필업", "재고없음", "필업필요", "체크필요", "필업검토"]
            latest["status"] = np.select(cond, choices, default="정상")

        # 물류에 재고 있고 온라인(공홈) 부족·소진 임박 등 실무 필업 목적만 포함
        cond_a = (
            (latest["daily_sales_7d"] > 0)
            & (latest["days_until_out"] < 14)
            & (latest["warehouse_stock"] > 0)
        )
        cond_b = (
            ((latest["stock"] <= 0) | (latest["online_stock"] <= 0))
            & (latest["warehouse_stock"] > 0)
            & (latest["daily_sales_7d"] > 0)
        )
        cond_c = (
            (latest["warehouse_stock"] > 0)
            & (latest["daily_sales_7d"] > 0)
            & (latest["days_until_out"] < 60)
        )
        df_target = latest[cond_a | cond_b | cond_c].copy()

        if "online_stock" not in df_target.columns:
            df_target["online_stock"] = (
                df_target["stock"] - df_target["channel_stock"]
            ).clip(lower=0).astype(int)

        df_target["_sort_days"] = df_target["days_until_out"].clip(upper=998)
        df_target = df_target.sort_values(
            ["_sort_days", "daily_sales_7d"],
            ascending=[True, False],
        ).drop(columns=["_sort_days"])

        if df_target.empty:
            flash("필업 대상 SKU가 없습니다.", "info")
            return redirect(url_for("dashboard"))

        output = _build_fillup_excel(df_target, str(latest_date))
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"필업지_{latest_date}.xlsx",
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        flash(f"필업지 생성 실패: {e}", "danger")
        return redirect(url_for("dashboard"))
    finally:
        conn.close()


@app.get("/export/current")
@login_required
def export_current():
    """현재 대시보드 데이터를 엑셀로 내보내기"""
    try:
        conn = get_conn()
        latest_date, latest = load_latest(conn)
        
        if latest_date is None or latest.empty:
            flash("내보낼 데이터가 없습니다.", "warning")
            return redirect(url_for("dashboard"))
        
        # 파일명 생성
        filename = f"재고현황_{latest_date}.xlsx"
        
        # 엑셀 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            latest.to_excel(writer, index=False, sheet_name='재고현황')
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f"엑셀 내보내기 실패: {e}", "danger")
        import traceback
        traceback.print_exc()
        return redirect(url_for("dashboard"))


@app.get("/export/database")
@login_required
def export_database():
    """전체 데이터베이스 백업"""
    from pathlib import Path
    
    try:
        db_path = Path(__file__).parent / "inventory.db"
        
        if not db_path.exists():
            flash("데이터베이스 파일이 없습니다.", "warning")
            return redirect(url_for("dashboard"))
        
        # 현재 날짜시간으로 파일명 생성
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventory_backup_{timestamp}.db"
        
        return send_file(
            db_path,
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f"DB 백업 실패: {e}", "danger")
        import traceback
        traceback.print_exc()
        return redirect(url_for("dashboard"))


@app.get("/change_password")
@login_required
def change_password_get():
    """비밀번호 변경 화면"""
    return render_template("change_password.html", title=APP_TITLE)


@app.post("/change_password")
@login_required
def change_password_post():
    """비밀번호 변경 처리"""
    current_pw = (request.form.get("current_password") or "").strip()
    new_pw = (request.form.get("new_password") or "").strip()
    confirm_pw = (request.form.get("confirm_password") or "").strip()
    
    expected = _expected_password()
    if not current_pw or current_pw != expected:
        flash("현재 비밀번호가 올바르지 않습니다.", "danger")
        return redirect(url_for("change_password_get"))
    
    if not new_pw:
        flash("신규 비밀번호를 입력하세요.", "danger")
        return redirect(url_for("change_password_get"))
    
    if new_pw != confirm_pw:
        flash("신규 비밀번호와 확인용 비밀번호가 일치하지 않습니다.", "danger")
        return redirect(url_for("change_password_get"))
    
    try:
        _set_password_in_db(new_pw)
        flash("비밀번호가 성공적으로 변경되었습니다.", "success")
        return redirect(url_for("dashboard"))
    except Exception as e:
        flash(f"비밀번호 변경에 실패했습니다: {e}", "danger")
        return redirect(url_for("change_password_get"))


@app.get("/")
def root():
    """루트 경로 리다이렉트"""
    return redirect(url_for("dashboard"))


@app.post("/settings/gsheet")
@login_required
def settings_gsheet_save():
    """구글 시트 연동 URL 설정 저장 (일반 폼 요청 + AJAX 공용)"""
    from flask import jsonify

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    gsheet_url = (request.form.get("gsheet_url") or "").strip()
    gsheet_sheet = (request.form.get("gsheet_sheet") or "재고판매현황").strip() or "재고판매현황"

    if not gsheet_url:
        if is_ajax:
            return jsonify(ok=False, error="URL 또는 ID를 입력하세요."), 400
        flash("스프레드시트 URL 또는 ID를 입력하세요.", "danger")
        return redirect(url_for("upload_get"))

    try:
        _set_gsheet_settings(gsheet_url, gsheet_sheet)
        if is_ajax:
            return jsonify(ok=True)
        flash("✅ 구글 시트 연동 설정이 저장되었습니다.", "success")
    except Exception as e:
        if is_ajax:
            return jsonify(ok=False, error=str(e)), 500
        flash(f"설정 저장 실패: {e}", "danger")

    return redirect(url_for("upload_get"))


@app.post("/sync/gsheet")
@login_required
def gsheet_sync_post():
    """저장된 구글 시트에서 1번(상품분석판매)만 읽어 해당 스냅샷 날짜로 DB 갱신"""
    saved = _get_gsheet_settings()
    url = (saved.get("url") or "").strip()
    sheet = (saved.get("sheet") or "재고판매현황").strip() or "재고판매현황"
    if not url:
        flash("구글 시트 연동 URL이 없습니다. 업로드 페이지에서 먼저 저장하세요.", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    snap_raw = (request.form.get("snapshot_date") or "").strip()
    try:
        date = dt.date.fromisoformat(snap_raw) if snap_raw else dt.date.today()
    except ValueError:
        flash("날짜 형식이 올바르지 않습니다(YYYY-MM-DD).", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    ref = request.referrer or url_for("dashboard")
    try:
        conn = get_conn()
        sales_df = fetch_gsheet_as_dataframe(url, sheet_name=sheet)
        print(f"[INFO] DB 최신화(구글시트): {len(sales_df)}행")

        result = normalize_excel(sales_df, snapshot_date=date, return_failed=True)
        if isinstance(result, tuple):
            sales_snap, failed_df = result
            if not failed_df.empty:
                failed_csv_path = f"failed_upload_{date.isoformat()}.csv"
                failed_df.to_csv(failed_csv_path, index=False, encoding="utf-8-sig")
                session["failed_csv_path"] = failed_csv_path
                session["failed_count"] = len(failed_df)
            else:
                session.pop("failed_csv_path", None)
                session.pop("failed_count", None)
        else:
            sales_snap = result
            session.pop("failed_csv_path", None)
            session.pop("failed_count", None)

        sales_count = upsert_snapshot(conn, sales_snap)
        _invalidate_snapshot_caches()
        flash(
            f"✅ DB 최신화 완료: 구글 시트 → {sales_count}개 품목 반영 (기준일 {date})",
            "success",
        )
        if session.get("failed_count"):
            flash(
                f"⚠️ {session['failed_count']}개 행은 반영되지 않았습니다. 실패 목록을 다운로드하세요.",
                "warning",
            )
    except Exception as e:
        flash(f"DB 최신화 실패: {e}", "danger")
        import traceback
        traceback.print_exc()

    return redirect(ref)


@app.get("/upload")
@login_required
def upload_get():
    """업로드 페이지"""
    gsheet_settings = _get_gsheet_settings()
    return render_template(
        "upload.html",
        title=APP_TITLE,
        default_date=dt.date.today().isoformat(),
        gsheet_settings=gsheet_settings,
    )


@app.post("/upload")
@login_required
def upload_post():
    """파일 업로드 처리"""
    sales_source = (request.form.get("sales_source") or "file").strip()  # "file" | "gsheet"
    sales_file = request.files.get("sales_file")
    warehouse_file = request.files.get("warehouse_file")
    warehouse_file2 = request.files.get("warehouse_file2")
    channel_file = request.files.get("channel_file")
    distribution_file = request.files.get("distribution_file")
    omni_file = request.files.get("omni_file")
    snapshot_date = (request.form.get("snapshot_date") or "").strip()
    warehouse_sheet = (request.form.get("warehouse_sheet") or "").strip()
    warehouse2_sheet = (request.form.get("warehouse2_sheet") or "").strip()
    channel_sheet = (request.form.get("channel_sheet") or "").strip()
    distribution_sheet = (request.form.get("distribution_sheet") or "").strip()
    omni_sheet = (request.form.get("omni_sheet") or "").strip()
    # 구글 시트 전용 파라미터
    gsheet_input = (request.form.get("gsheet_input") or "").strip()
    gsheet_sheet = (request.form.get("gsheet_sheet") or "재고판매현황").strip()

    # 구글 시트 소스: 입력이 없으면 DB 저장값 자동 사용
    if sales_source == "gsheet" and not gsheet_input:
        saved = _get_gsheet_settings()
        gsheet_input = saved.get("url", "")
        if not gsheet_sheet or gsheet_sheet == "재고판매현황":
            gsheet_sheet = saved.get("sheet", "재고판매현황")

    # 필수 소스 검증
    if sales_source == "gsheet":
        if not gsheet_input:
            flash("구글 시트 URL을 확인할 수 없습니다. 업로드 페이지에서 주소를 저장하거나 다시 시도해 주세요.", "danger")
            return redirect(url_for("upload_get"))
    else:
        if not sales_file or not sales_file.filename:
            flash("상품분석판매 파일을 선택하세요.", "danger")
            return redirect(url_for("upload_get"))

    # 날짜 파싱
    try:
        date = dt.date.fromisoformat(snapshot_date) if snapshot_date else dt.date.today()
    except ValueError:
        flash("날짜 형식이 올바르지 않습니다(YYYY-MM-DD).", "danger")
        return redirect(url_for("upload_get"))

    conn = None
    try:
        conn = get_conn()

        # 1. 상품분석판매 로드 (파일 업로드 또는 구글 시트)
        if sales_source == "gsheet":
            try:
                sales_df = fetch_gsheet_as_dataframe(gsheet_input, sheet_name=gsheet_sheet)
                print(f"[INFO] 구글 시트 로드 완료: {len(sales_df)}행, 컬럼: {sales_df.columns.tolist()}")
            except Exception as gs_err:
                flash(f"구글 시트 읽기 실패: {gs_err}", "danger")
                return redirect(url_for("upload_get"))
        else:
            sales_filename = sales_file.filename or ""
            sales_ext = os.path.splitext(sales_filename)[1].lower()

            if sales_ext == ".csv":
                try:
                    sales_df = pd.read_csv(sales_file)
                except UnicodeDecodeError:
                    sales_file.seek(0)
                    sales_df = pd.read_csv(sales_file, encoding="cp949")
            elif sales_ext in [".xlsx", ".xls", ".xlsb"]:
                sales_df = pd.read_excel(sales_file, sheet_name=0)
            else:
                flash("지원하지 않는 파일 형식입니다. CSV 또는 Excel 파일을 사용하세요.", "danger")
                return redirect(url_for("upload_get"))

        # return_failed=True로 호출하여 실패한 행도 받기
        result = normalize_excel(sales_df, snapshot_date=date, return_failed=True)
        if isinstance(result, tuple):
            sales_snap, failed_df = result
            # 실패한 행이 있으면 CSV로 저장
            if not failed_df.empty:
                failed_csv_path = f"failed_upload_{date.isoformat()}.csv"
                failed_df.to_csv(failed_csv_path, index=False, encoding='utf-8-sig')
                session['failed_csv_path'] = failed_csv_path
                session['failed_count'] = len(failed_df)
        else:
            sales_snap = result
        
        sales_count = upsert_snapshot(conn, sales_snap)
        conn.commit()

        # 2. 물류센터1 재고 업로드 (선택사항)
        warehouse1_count = 0
        if warehouse_file and warehouse_file.filename:
            warehouse_filename = warehouse_file.filename or ""
            warehouse_ext = os.path.splitext(warehouse_filename)[1].lower()
            
            if warehouse_ext in [".xlsx", ".xls", ".xlsb"]:
                warehouse_df = pd.read_excel(warehouse_file, sheet_name=(warehouse_sheet or 0))
                warehouse_df.columns = [str(c).strip() for c in warehouse_df.columns]
                print(f"[INFO] 물류센터1 엑셀 컬럼: {warehouse_df.columns.tolist()}")
                print(f"[INFO] 물류센터1 엑셀 행 수: {len(warehouse_df)}")

                if warehouse_df.empty:
                    flash("⚠️ 항만 물류센터 파일에서 유효한 데이터를 찾을 수 없습니다.", "warning")
                else:
                    sku_warehouse_map = {}
                    sku_solid_map = {}
                    sku_assort_map = {}
                    sku_ratio_map = {}
                    sku_box_map = {}

                    def _wh_int(val, default=0) -> int:
                        try:
                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                return default
                            return int(float(val))
                        except (TypeError, ValueError):
                            return default

                    for _, row in warehouse_df.iterrows():
                        sku = str(row.get("상품", "") or "").strip()
                        if not sku or len(sku) != 15:
                            continue

                        solid = _wh_int(row.get("솔리드가용재고", 0))
                        assort_raw = row.get("아소트 가용재고", row.get("아소트가용재고", 0))
                        assort = _wh_int(assort_raw)
                        ratio = _wh_int(row.get("아소트비율", 0))
                        box_count = _wh_int(row.get("아소트박스수", 0))

                        total = solid + assort
                        sku_warehouse_map[sku] = total
                        sku_solid_map[sku] = solid
                        sku_assort_map[sku] = assort
                        if ratio > 0:
                            sku_ratio_map[sku] = ratio
                        if box_count > 0:
                            sku_box_map[sku] = box_count

                    if sku_warehouse_map:
                        warehouse1_count = update_warehouse_stock(
                            conn,
                            date.isoformat(),
                            sku_warehouse_map,
                            warehouse_num=1,
                            solid_map=sku_solid_map if sku_solid_map else None,
                            assort_map=sku_assort_map if sku_assort_map else None,
                            assort_ratio_map=sku_ratio_map if sku_ratio_map else None,
                            assort_box_map=sku_box_map if sku_box_map else None,
                        )
                        conn.commit()
                        total_warehouse = sum(sku_warehouse_map.values())
                        print(f"[INFO] 물류센터1 업로드: {len(sku_warehouse_map)}개 SKU, 총 재고: {total_warehouse}")
                        print(f"[INFO] 업데이트된 SKU: {warehouse1_count}개")
                    else:
                        flash("⚠️ 항만 물류센터 파일에서 15자리 SKU를 찾을 수 없습니다.", "warning")
            else:
                flash("항만 물류센터: Excel 파일만 지원됩니다.", "warning")
        
        # 3. 물류센터2 재고 업로드 (선택사항)
        warehouse2_count = 0
        if warehouse_file2 and warehouse_file2.filename:
            warehouse2_filename = warehouse_file2.filename or ""
            warehouse2_ext = os.path.splitext(warehouse2_filename)[1].lower()
            
            if warehouse2_ext in [".xlsx", ".xls", ".xlsb"]:
                warehouse2_df = pd.read_excel(warehouse_file2, sheet_name=(warehouse2_sheet or 0))
                warehouse2_df.columns = [str(c).strip() for c in warehouse2_df.columns]
                print(f"[INFO] 물류센터2 엑셀 컬럼: {warehouse2_df.columns.tolist()}")
                print(f"[INFO] 물류센터2 엑셀 행 수: {len(warehouse2_df)}")

                if warehouse2_df.empty:
                    flash("⚠️ 부평 물류센터 파일에서 유효한 데이터를 찾을 수 없습니다.", "warning")
                else:
                    sku_warehouse2_map = {}
                    sku2_solid_map = {}
                    sku2_assort_map = {}
                    sku2_ratio_map = {}
                    sku2_box_map = {}

                    def _wh2_int(val, default=0) -> int:
                        try:
                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                return default
                            return int(float(val))
                        except (TypeError, ValueError):
                            return default

                    for _, row in warehouse2_df.iterrows():
                        sku = str(row.get("상품", "") or "").strip()
                        if not sku or len(sku) != 15:
                            continue

                        solid = _wh2_int(row.get("솔리드가용재고", 0))
                        assort_raw = row.get("아소트 가용재고", row.get("아소트가용재고", 0))
                        assort = _wh2_int(assort_raw)
                        ratio = _wh2_int(row.get("아소트비율", 0))
                        box_count = _wh2_int(row.get("아소트박스수", 0))

                        total = solid + assort
                        sku_warehouse2_map[sku] = total
                        sku2_solid_map[sku] = solid
                        sku2_assort_map[sku] = assort
                        if ratio > 0:
                            sku2_ratio_map[sku] = ratio
                        if box_count > 0:
                            sku2_box_map[sku] = box_count

                    if sku_warehouse2_map:
                        warehouse2_count = update_warehouse_stock(
                            conn,
                            date.isoformat(),
                            sku_warehouse2_map,
                            warehouse_num=2,
                            solid_map=sku2_solid_map if sku2_solid_map else None,
                            assort_map=sku2_assort_map if sku2_assort_map else None,
                            assort_ratio_map=sku2_ratio_map if sku2_ratio_map else None,
                            assort_box_map=sku2_box_map if sku2_box_map else None,
                        )
                        conn.commit()
                        total_warehouse2 = sum(sku_warehouse2_map.values())
                        print(f"[INFO] 물류센터2 업로드: {len(sku_warehouse2_map)}개 SKU, 총 재고: {total_warehouse2}")
                        print(f"[INFO] 업데이트된 SKU: {warehouse2_count}개")
                    else:
                        flash("⚠️ 부평 물류센터 파일에서 15자리 SKU를 찾을 수 없습니다.", "warning")
            else:
                flash("부평 물류센터: Excel 파일만 지원됩니다.", "warning")
        
        # 4. 매장 재고 업로드 (선택사항)
        channel_count = 0
        if channel_file and channel_file.filename:
            channel_filename = channel_file.filename or ""
            channel_ext = os.path.splitext(channel_filename)[1].lower()
            
            if channel_ext in [".xlsx", ".xls", ".xlsb"]:
                channel_df = pd.read_excel(channel_file, sheet_name=(channel_sheet or 0))
                print(f"[INFO] 매장재고 엑셀 컬럼: {channel_df.columns.tolist()}")
                print(f"[INFO] 매장재고 엑셀 행 수: {len(channel_df)}")
                
                channel_snap = normalize_excel(channel_df, snapshot_date=date)
                print(f"[INFO] 정규화 후 행 수: {len(channel_snap)}")
                
                if channel_snap.empty:
                    flash("⚠️ 매장재고 파일에서 유효한 데이터를 찾을 수 없습니다.", "warning")
                else:
                    # SKU와 매장재고 매핑
                    sku_channel_map = {}
                    for _, row in channel_snap.iterrows():
                        sku = str(row["sku"]).strip()
                        channel_stock = int(row.get("channel_stock") or 0)
                        if sku and len(sku) == 15:
                            sku_channel_map[sku] = channel_stock
                    
                    if sku_channel_map:
                        channel_count = update_channel_stock(
                            conn, date.isoformat(), sku_channel_map
                        )
                        conn.commit()
                        total_channel = sum(sku_channel_map.values())
                        print(f"[INFO] 매장재고 업로드: {len(sku_channel_map)}개 SKU, 총 재고: {total_channel}")
                        print(f"[INFO] 업데이트된 SKU: {channel_count}개")
                    else:
                        flash("⚠️ 매장재고 파일에서 15자리 SKU를 찾을 수 없습니다.", "warning")
            else:
                flash("매장재고: Excel 파일만 지원됩니다.", "warning")
        
        # 5. 분배내역 업로드 (선택사항)
        distribution_count = 0
        if distribution_file and distribution_file.filename:
            distribution_filename = distribution_file.filename or ""
            distribution_ext = os.path.splitext(distribution_filename)[1].lower()
            if distribution_ext in [".xlsx", ".xls", ".xlsb"]:
                try:
                    dist_df = pd.read_excel(distribution_file, sheet_name=(distribution_sheet or 0))
                    dist_df.columns = [str(c).strip() for c in dist_df.columns]
                    # SKU 컬럼 후보: SKU, 상품코드, 상품, 품목코드
                    sku_col = None
                    for col in ["SKU", "상품코드", "상품", "품목코드", "sku"]:
                        if col in dist_df.columns:
                            sku_col = col
                            break
                    # 분배량 컬럼 우선 (N열 등): 분배량, 수량, N열 → 수량 합계로 표시
                    qty_col = None
                    for col in ["분배량", "수량", "분배수량"]:
                        if col in dist_df.columns:
                            qty_col = col
                            break
                    if not qty_col and len(dist_df.columns) >= 14:
                        # N열 = 14번째 컬럼(인덱스 13)
                        qty_col = dist_df.columns[13]
                    # 텍스트 비고 컬럼 (분배량 없을 때 대체)
                    note_col = None
                    for col in ["분배내역", "비고", "메모", "내역", "분배요청내역", "분배요청", "비고사항"]:
                        if col in dist_df.columns:
                            note_col = col
                            break
                    use_qty = qty_col is not None
                    use_note = note_col is not None and not use_qty
                    if sku_col and (use_qty or use_note):
                        sku_note_map = {}
                        for _, row in dist_df.iterrows():
                            sku_raw = str(row.get(sku_col, "")).strip()
                            sku = sku_raw[:15] if len(sku_raw) >= 15 else sku_raw
                            if not sku or sku == "nan":
                                continue
                            if use_qty:
                                val = row.get(qty_col)
                                qty = int(pd.to_numeric(val, errors="coerce")) if not pd.isna(val) else 0
                                if sku in sku_note_map:
                                    sku_note_map[sku] = sku_note_map[sku] + qty
                                else:
                                    sku_note_map[sku] = qty
                            else:
                                note_val = row.get(note_col)
                                note = "" if pd.isna(note_val) else str(note_val).strip()
                                if sku in sku_note_map:
                                    sku_note_map[sku] = sku_note_map[sku] + " / " + note
                                else:
                                    sku_note_map[sku] = note
                        if use_qty:
                            sku_note_map = {k: str(v) for k, v in sku_note_map.items()}
                        if sku_note_map:
                            distribution_count = update_distribution_note(
                                conn, date.isoformat(), sku_note_map
                            )
                            conn.commit()
                            print(f"[INFO] 분배내역 업로드: {distribution_count}개 SKU 반영 (분배량 기준)" if use_qty else f"[INFO] 분배내역 업로드: {distribution_count}개 SKU 반영")
                    else:
                        flash("⚠️ 분배내역 파일에 SKU(또는 상품코드) 컬럼과 분배량(또는 N열/수량) 컬럼이 필요합니다.", "warning")
                except Exception as ex:
                    flash(f"⚠️ 분배내역 파일 처리 중 오류: {ex}", "warning")
            else:
                flash("분배내역: Excel 파일만 지원됩니다.", "warning")
        
        # 6. 옴니판매불가 SKU 업로드 (선택사항)
        omni_count = 0
        if omni_file and omni_file.filename:
            omni_filename = omni_file.filename or ""
            omni_ext = os.path.splitext(omni_filename)[1].lower()
            if omni_ext in [".xlsx", ".xls", ".xlsb"]:
                try:
                    sheet = omni_sheet or 0
                    read_kwargs = {"sheet_name": sheet}
                    if omni_ext == ".xls":
                        try:
                            import xlrd  # noqa: F401
                            read_kwargs["engine"] = "xlrd"
                        except ImportError:
                            raise ImportError(
                                "옴니판매불가 .xls 파일을 읽으려면 xlrd 패키지가 필요합니다. "
                                "pip install xlrd 후 다시 시도하거나, 엑셀에서 .xlsx 형식으로 저장해 주세요."
                            )
                    omni_df = pd.read_excel(omni_file, **read_kwargs)
                    
                    if omni_df.shape[1] < 8:
                        flash("⚠️ 옴니판매불가 파일에 필요한 열(C,D,E,H)이 부족합니다.", "warning")
                    else:
                        df = omni_df.copy()
                        # C열=매장명, D열=스타일코드, E열=단품코드, H열=판매불가 수량
                        store_col = df.columns[2]
                        style_col = df.columns[3]
                        sku_col = df.columns[4]
                        blocked_col = df.columns[7]
                        
                        df["store_name"] = df[store_col].astype(str).str.strip()
                        df["style_code"] = df[style_col].astype(str).str.strip()
                        df["sku_code"] = df[sku_col].astype(str).str.strip()
                        df["blocked_qty"] = (
                            pd.to_numeric(df[blocked_col], errors="coerce")
                            .fillna(0)
                            .astype(int)
                        )
                        
                        df = df[
                            (df["style_code"] != "")
                            & (df["sku_code"] != "")
                            & (df["blocked_qty"] > 0)
                        ].copy()
                        
                        if df.empty:
                            flash("⚠️ 옴니판매불가 파일에서 유효한 데이터를 찾을 수 없습니다.", "warning")
                        else:
                            # 스타일/단품별 판매불가 수량 합계
                            summary = (
                                df.groupby(["style_code", "sku_code"], as_index=False)["blocked_qty"]
                                .sum()
                            )
                            
                            # 스타일/단품/매장별 합계 후, 각 쌍에서 가장 큰 매장 선택
                            store_agg = (
                                df.groupby(
                                    ["style_code", "sku_code", "store_name"], as_index=False
                                )["blocked_qty"]
                                .sum()
                            )
                            store_agg = store_agg.sort_values(
                                ["style_code", "sku_code", "blocked_qty"],
                                ascending=[True, True, False],
                            )
                            top_store_df = store_agg.drop_duplicates(
                                subset=["style_code", "sku_code"], keep="first"
                            ).rename(
                                columns={"store_name": "top_store"}
                            )
                            
                            omni_join = summary.merge(
                                top_store_df[["style_code", "sku_code", "top_store"]],
                                on=["style_code", "sku_code"],
                                how="left",
                            )
                            
                            # 기존 데이터 삭제 후 삽입
                            conn.execute(
                                "DELETE FROM omni_blocked WHERE snapshot_date = ?",
                                (date.isoformat(),),
                            )
                            rows = [
                                (
                                    date.isoformat(),
                                    str(r["style_code"]),
                                    str(r["sku_code"]),
                                    int(r["blocked_qty"]),
                                    str(r.get("top_store") or ""),
                                )
                                for _, r in omni_join.iterrows()
                            ]
                            conn.executemany(
                                """
                                INSERT INTO omni_blocked (
                                    snapshot_date, style_code, sku_code, blocked_qty, top_store
                                ) VALUES (?, ?, ?, ?, ?)
                                """,
                                rows,
                            )
                            conn.commit()
                            omni_count = len(rows)
                            print(f"[INFO] 옴니판매불가 업로드: {omni_count}개 단품")
                except Exception as ex:
                    flash(f"⚠️ 옴니판매불가 파일 처리 중 오류: {ex}", "warning")
            else:
                flash("옴니판매불가: Excel 파일만 지원됩니다.", "warning")
        
        # 결과 메시지
        total_warehouse_count = warehouse1_count + warehouse2_count
        msg_parts = [f"상품분석판매: {sales_count}개 품목"]
        if warehouse1_count > 0:
            msg_parts.append(f"항만 물류센터: {warehouse1_count}개 SKU")
        if warehouse2_count > 0:
            msg_parts.append(f"부평 물류센터: {warehouse2_count}개 SKU")
        if channel_count > 0:
            msg_parts.append(f"매장재고: {channel_count}개 SKU")
        if distribution_count > 0:
            msg_parts.append(f"분배내역: {distribution_count}개 SKU")
        if 'omni_count' in locals() and omni_count > 0:
            msg_parts.append(f"옴니판매불가: {omni_count}개 단품")
        
        success_msg = f"✅ {', '.join(msg_parts)} 업로드 완료 (날짜: {date})"
        flash(success_msg, "success")
        _invalidate_snapshot_caches()

        # 실패한 행이 있으면 알림 (다운로드는 상단 배너에서 가능)
        if 'failed_count' in session and session['failed_count'] > 0:
            flash(f"⚠️ {session['failed_count']}개 행이 업로드 실패했습니다. 상단 배너에서 실패 목록을 다운로드하세요.", "warning")
        
        return redirect(url_for("dashboard"))

    except Exception as e:
        flash(f"업로드 실패: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for("upload_get"))
    finally:
        if conn:
            conn.close()


def _status_badge(status: str) -> str:
    """상태별 Bootstrap 색상 클래스 반환"""
    status_colors = {
        "긴급필업": "danger",
        "재고없음": "dark",
        "필업필요": "warning",
        "체크필요": "info",
        "저재고": "warning",
        "필업검토": "secondary",
        "정상": "success",
    }
    return status_colors.get(status, "secondary")


@app.get("/dashboard")
@login_required
@cache.cached(timeout=300, query_string=True, key_prefix="dashboard")
def dashboard():
    # [최적화] GET+쿼리스트링별 300초 캐시; 동기화/업로드/초기화 시 _invalidate_snapshot_caches()
    """대시보드 메인 화면"""
    try:
        return _dashboard_impl()
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[ERROR] 대시보드 오류: {e}")
        print(tb)
        return (
            "<h1>500 Internal Server Error</h1>"
            "<pre style='background:#fdd;padding:1em;overflow:auto;'>"
            + tb.replace("<", "&lt;").replace(">", "&gt;")
            + "</pre>",
            500,
        )


def _item_code_from_sku(sku) -> str:
    """스타일코드 10자리(=SKU 앞 10자) 기준 3·4번째 문자 → 예: SPJPG11C24 → JP"""
    s = str(sku).strip()
    if len(s) < 4:
        return ""
    return s[2:4].upper()


def _season_code_sort_key_item(c: str) -> tuple:
    s = str(c).upper().strip()
    if len(s) < 2:
        return (2, s)
    rest = s[1:]
    if rest.isdigit():
        return (0, int(rest))
    return (1, rest)


def _df_filter_item_tab(
    df: pd.DataFrame, current_season_letter: str, item_tab: str
) -> pd.DataFrame:
    """올해 시즌(G*) 행만, item_tab이 all이 아니면 해당 시즌코드만."""
    if df.empty or "sku" not in df.columns:
        return df.iloc[0:0].copy()
    sc = df["sku"].astype(str).str[4:6].str.strip().str.upper()
    mask = sc.str.startswith(current_season_letter) & (sc.str.len() >= 2)
    tab = (item_tab or "all").strip().upper()
    if tab and tab != "ALL":
        mask = mask & (sc == tab)
    return df.loc[mask].copy()


def _sql_prev_item_stock_totals(
    conn,
    prev_date: str,
    current_season_letter: str,
    item_tab: str,
) -> pd.DataFrame:
    """직전 스냅샷에서 아이템코드(SKU 3~4자리)별 재고 합. `_df_filter_item_tab`과 동일한 시즌/탭 필터."""
    tab_u = (item_tab or "all").strip().upper()
    letter = (current_season_letter or "").strip().upper()[:1]
    if not letter:
        return pd.DataFrame(columns=["item_code", "stock_prev"])
    sql = """
        SELECT
            UPPER(SUBSTR(TRIM(sku), 3, 2)) AS item_code,
            SUM(COALESCE(stock, 0)) AS stock_prev
        FROM snapshots
        WHERE snapshot_date = ?
          AND LENGTH(TRIM(sku)) >= 6
          AND UPPER(SUBSTR(TRIM(sku), 5, 1)) = ?
          AND LENGTH(TRIM(SUBSTR(TRIM(sku), 5, 2))) >= 2
    """
    params: list = [prev_date, letter]
    if tab_u and tab_u != "ALL":
        sql += " AND UPPER(SUBSTR(TRIM(sku), 5, 2)) = ? "
        params.append(tab_u)
    sql += """
        GROUP BY UPPER(SUBSTR(TRIM(sku), 3, 2))
        HAVING LENGTH(TRIM(item_code)) > 0
    """
    return pd.read_sql_query(sql, conn, params=params)


def _build_item_inventory_summary(
    conn,
    latest_date: str,
    latest_df: pd.DataFrame,
    current_season_letter: str,
    item_tab: str,
    allowed_season_tabs: list[str],
) -> tuple[list[dict], Optional[str], bool]:
    """
    최신 스냅샷 기준 아이템별 총재고·총판매량·판매량 비중, 직전 스냅샷 대비 재고 증감.
    테이블 집계: 올해 시즌 + (탭: 전체 또는 G1/G2/…).
    상세(offcanvas): 올해 시즌만, 시즌별 행은 G1~GA 전부.
    정렬: 총 판매량 내림차순.
    Returns: (rows, prev_date or None, has_prev)
    """
    cache_key = f"item_summary_{latest_date}_{item_tab}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    dates_df = pd.read_sql_query(
        """
        SELECT DISTINCT snapshot_date AS d
        FROM snapshots
        ORDER BY snapshot_date DESC
        LIMIT 2
        """,
        conn,
    )
    if dates_df.empty:
        return [], None, False
    prev_date: Optional[str] = None
    if len(dates_df) >= 2:
        prev_date = str(dates_df.iloc[1]["d"])
    has_prev = prev_date is not None

    tab_u = (item_tab or "all").strip().upper()
    if tab_u != "ALL" and tab_u not in {s.upper() for s in (allowed_season_tabs or [])}:
        tab_u = "ALL"

    def prep_work(df: pd.DataFrame) -> pd.DataFrame:
        work = df.copy()
        if work.empty:
            return pd.DataFrame(
                columns=["item_code", "sku", "stock", "sales_qty", "season_code", "is_oos", "name"]
            )
        work["item_code"] = work["sku"].map(_item_code_from_sku)
        work = work[work["item_code"] != ""]
        if "sales_qty" not in work.columns:
            work["sales_qty"] = 0
        work["sales_qty"] = pd.to_numeric(work["sales_qty"], errors="coerce").fillna(0)
        work["stock"] = pd.to_numeric(work["stock"], errors="coerce").fillna(0)
        if "name" not in work.columns:
            work["name"] = ""
        work["name"] = work["name"].fillna("").astype(str)
        work["sku"] = work["sku"].astype(str)
        work["season_code"] = work["sku"].str[4:6].str.upper()
        work["is_oos"] = (work["stock"] <= 0).astype(int)
        return work

    def agg_items(df: pd.DataFrame) -> pd.DataFrame:
        work = prep_work(df)
        if work.empty:
            return pd.DataFrame(
                columns=["item_code", "total_stock", "total_sales", "sku_total", "sku_oos", "oos_rate"]
            )
        g = work.groupby("item_code", as_index=False).agg(
            total_stock=("stock", "sum"),
            total_sales=("sales_qty", "sum"),
            sku_total=("sku", "nunique"),
            sku_oos=("is_oos", "sum"),
        )
        g["oos_rate"] = (
            (g["sku_oos"] / g["sku_total"] * 100.0).fillna(0).round(1)
            if not g.empty
            else 0.0
        )
        return g

    latest_for_table = _df_filter_item_tab(latest_df, current_season_letter, tab_u)
    cur_agg = agg_items(latest_for_table)
    if cur_agg.empty:
        return [], prev_date, has_prev

    total_sales_all = float(cur_agg["total_sales"].sum())
    cur_agg["sales_share_pct"] = 0.0
    if total_sales_all > 0:
        cur_agg["sales_share_pct"] = (cur_agg["total_sales"] / total_sales_all * 100.0).round(2)

    if has_prev and prev_date:
        prev_agg = _sql_prev_item_stock_totals(conn, prev_date, current_season_letter, tab_u)
        merged = cur_agg.merge(prev_agg, on="item_code", how="left")
        merged["stock_prev"] = merged["stock_prev"].fillna(0).astype(int)
        merged["stock_delta"] = merged["total_stock"].astype(int) - merged["stock_prev"]
    else:
        merged = cur_agg.copy()
        merged["stock_prev"] = pd.NA
        merged["stock_delta"] = pd.NA

    merged = merged.sort_values("total_sales", ascending=False)

    # 상세용: 올해 시즌 문자로 시작하는 행만 (탭과 무관)
    latest_work = prep_work(latest_df)
    scy = latest_work["season_code"].astype(str).str.strip().str.upper()
    mask_y = scy.str.startswith(current_season_letter) & (scy.str.len() >= 2)
    latest_work_year = latest_work[mask_y].copy()

    season_top_map: dict[str, list[dict]] = {}
    item_oos_top20_map: dict[str, list[dict]] = {}
    item_imminent_top20_map: dict[str, list[dict]] = {}
    if not latest_work_year.empty:
        season_stat = (
            latest_work_year.groupby(["item_code", "season_code"], as_index=False)
            .agg(total=("sku", "nunique"), stockout=("is_oos", "sum"))
        )
        season_stat["rate"] = (season_stat["stockout"] / season_stat["total"] * 100.0).fillna(0).round(1)
        for item_code, grp in season_stat.groupby("item_code"):
            grp_sorted = grp.copy()
            grp_sorted["_sk"] = grp_sorted["season_code"].astype(str).map(_season_code_sort_key_item)
            grp_sorted = grp_sorted.sort_values("_sk").drop(columns=["_sk"])
            season_top_map[str(item_code)] = [
                {
                    "code": str(rr["season_code"]),
                    "rate": float(rr["rate"]),
                    "stockout": int(rr["stockout"]),
                    "total": int(rr["total"]),
                }
                for _, rr in grp_sorted.iterrows()
            ]

        # 아이템 기준 결품 SKU Top20: 재고<=0, 판매량 높은 순
        oos_candidates = latest_work_year[latest_work_year["is_oos"] == 1].copy()
        if not oos_candidates.empty:
            oos_candidates = oos_candidates.sort_values(
                ["item_code", "sales_qty", "sku"], ascending=[True, False, True]
            )
            for item_code, grp in oos_candidates.groupby("item_code"):
                top20 = grp.head(20)
                item_oos_top20_map[str(item_code)] = [
                    {
                        "sku": str(rr["sku"]),
                        "name": str(rr.get("name") or ""),
                        "sales_qty": int(rr.get("sales_qty") or 0),
                        "stock": int(rr.get("stock") or 0),
                        "season_code": str(rr.get("season_code") or ""),
                    }
                    for _, rr in top20.iterrows()
                ]

        # 결품임박 Top20: 재고>0 이고 판매량>0, (재고÷일판매) 낮은 순 = 빨리 소진
        im = latest_work_year[
            (latest_work_year["stock"] > 0) & (latest_work_year["sales_qty"] > 0)
        ].copy()
        if not im.empty:
            daily = im["sales_qty"].astype(float) / 7.0
            im = im.assign(_daily=daily)
            im["_cover"] = im["stock"].astype(float) / im["_daily"].replace(0, float("nan"))
            im = im.sort_values(
                ["item_code", "_cover", "sales_qty", "sku"],
                ascending=[True, True, False, True],
                na_position="last",
            )
            for item_code, grp in im.groupby("item_code"):
                top20i = grp.head(20)
                item_imminent_top20_map[str(item_code)] = [
                    {
                        "sku": str(rr["sku"]),
                        "name": str(rr.get("name") or ""),
                        "sales_qty": int(rr.get("sales_qty") or 0),
                        "stock": int(rr.get("stock") or 0),
                        "season_code": str(rr.get("season_code") or ""),
                    }
                    for _, rr in top20i.iterrows()
                ]

    rows = []
    for _, r in merged.iterrows():
        sp = r["stock_prev"]
        sd = r["stock_delta"]
        rows.append(
            {
                "item_code": str(r["item_code"]),
                "total_stock": int(r["total_stock"]),
                "stock_prev": int(sp) if pd.notna(sp) else None,
                "stock_delta": int(sd) if pd.notna(sd) else None,
                "total_sales": int(r["total_sales"]),
                "sales_share_pct": float(r["sales_share_pct"]),
                "oos_rate": float(r["oos_rate"]),
                "season_oos_top": season_top_map.get(str(r["item_code"]), []),
                "item_oos_top20": item_oos_top20_map.get(str(r["item_code"]), []),
                "item_imminent_top20": item_imminent_top20_map.get(str(r["item_code"]), []),
            }
        )
    cache.set(cache_key, (rows, prev_date, has_prev), timeout=300)
    return rows, prev_date, has_prev


def _load_base_data() -> dict:
    # [최적화] 스냅샷 공통 dashboard_base; suggested_order_qty·긴급·item_tab 제외, conn 내부만 사용
    cached = cache.get("dashboard_base")
    if cached is not None:
        return cached
    import numpy as np

    conn = get_conn()
    try:
        latest_date, latest = load_latest(conn)
        if latest_date is None or latest.empty:
            return {"latest_date": None}

        all_data = latest.copy()
        for col in (
            "sales_qty",
            "channel_stock",
            "warehouse_stock",
            "warehouse1_stock",
            "warehouse2_stock",
            "min_stock",
            "lead_time_days",
            "safety_stock",
        ):
            if col not in all_data.columns:
                all_data[col] = 0
        if "distribution_note" not in all_data.columns:
            all_data["distribution_note"] = ""
        all_data["distribution_note"] = all_data["distribution_note"].fillna("").astype(str)
        all_data["category"] = all_data["category"].fillna("")
        all_data["season_code"] = all_data["sku"].astype(str).str[4:6]
        _current_year = dt.date.today().year
        current_season_letter = chr(ord("A") + (_current_year - 2020))
        prev_season_letter = chr(ord("A") + (_current_year - 2021))
        all_data["category_code"] = all_data["sku"].astype(str).str[7]
        all_data["sales_qty"] = all_data["sales_qty"].fillna(0).astype(int)
        all_data["daily_sales_7d"] = (all_data["sales_qty"] / 7.0).round(2)
        all_data["channel_stock"] = all_data["channel_stock"].fillna(0).astype(int)
        all_data["warehouse_stock"] = all_data["warehouse_stock"].fillna(0).astype(int)
        all_data["warehouse1_stock"] = all_data["warehouse1_stock"].fillna(0).astype(int)
        all_data["warehouse2_stock"] = all_data["warehouse2_stock"].fillna(0).astype(int)
        all_data["total_available"] = all_data["stock"] + all_data["warehouse_stock"]
        all_data["days_until_out"] = 999.0
        mask_has_sales = all_data["daily_sales_7d"] > 0
        all_data.loc[mask_has_sales, "days_until_out"] = (
            all_data.loc[mask_has_sales, "total_available"]
            / all_data.loc[mask_has_sales, "daily_sales_7d"]
        ).round(1)
        all_data.loc[(all_data["total_available"] == 0), "days_until_out"] = 0.0
        all_data["min_stock"] = all_data["min_stock"].fillna(0).astype(int)
        all_data["lead_time_days"] = all_data["lead_time_days"].fillna(7).astype(int)
        all_data["safety_stock"] = all_data["safety_stock"].fillna(0).astype(int)
        all_data["reorder_point"] = all_data["safety_stock"] + (
            all_data["daily_sales_7d"] * all_data["lead_time_days"]
        )
        conditions = [
            (all_data["stock"] == 0) & (all_data["daily_sales_7d"] > 0),
            (all_data["stock"] == 0),
            (all_data["daily_sales_7d"] > 0) & (all_data["days_until_out"] < 7),
            (all_data["stock"] <= 10) & (all_data["daily_sales_7d"] > 0),
            (all_data["stock"] < all_data["min_stock"]) & (all_data["min_stock"] > 0),
            (all_data["stock"] <= all_data["reorder_point"]) & (all_data["daily_sales_7d"] > 0),
        ]
        choices = ["긴급필업", "재고없음", "필업필요", "체크필요", "저재고", "필업검토"]
        all_data["status"] = np.select(conditions, choices, default="정상")
        all_data["product_code"] = all_data["sku"].astype(str).str[:10]

        total_items_all = int(all_data["sku"].nunique())
        total_stock_all = int(all_data["stock"].sum())
        oos_all = int((all_data["status"] == "긴급필업").sum())
        low_all = int((all_data["status"] == "체크필요").sum())
        has_channel_all = int((all_data["channel_stock"] > 0).sum())
        total_channel_stock_all = int(all_data["channel_stock"].sum())
        has_warehouse_all = int((all_data["warehouse_stock"] > 0).sum())
        total_warehouse_stock_all = int(all_data["warehouse_stock"].sum())
        _sc_norm = all_data["season_code"].astype(str).str.strip().str.upper()
        _mask_curr_year = _sc_norm.str.startswith(current_season_letter) & (_sc_norm.str.len() >= 2)
        _df_curr_year = all_data[_mask_curr_year]
        total_items_curr_year = int(_df_curr_year["sku"].nunique()) if len(_df_curr_year) > 0 else 0
        stockout_count_curr_year = int((_df_curr_year["stock"] == 0).sum()) if len(_df_curr_year) > 0 else 0
        stockout_rate_curr_year = (
            round((stockout_count_curr_year / total_items_curr_year * 100), 1)
            if total_items_curr_year > 0
            else 0.0
        )
        dist_note_filled = all_data["distribution_note"].fillna("").astype(str).str.strip() != ""
        distribution_items_all = int(dist_note_filled.sum())
        _dist_vals = all_data.loc[dist_note_filled, "distribution_note"]
        distribution_total_qty_all = int(
            pd.to_numeric(_dist_vals.astype(str).str.strip(), errors="coerce").fillna(0).astype(int).sum()
        )

        def _category_stats_rows(data: pd.DataFrame) -> list:
            sub = data[data["category_code"].notna()]
            if sub.empty:
                return []
            ag = sub.groupby("category_code", as_index=False).agg(
                total=("stock", "count"),
                stockout=("stock", lambda s: int((s == 0).sum())),
            )
            ag["rate"] = (ag["stockout"] / ag["total"] * 100).where(ag["total"] > 0, 0).round(1)
            ag = ag.rename(columns={"category_code": "code"})
            ag = ag.sort_values("code")
            return ag.to_dict(orient="records")

        def _season_code_sort_key(c: str) -> tuple:
            s = str(c).upper().strip()
            if len(s) < 2:
                return (2, s)
            rest = s[1:]
            if rest.isdigit():
                return (0, int(rest))
            return (1, rest)

        _curr_mask = _sc_norm.str.startswith(current_season_letter) & (_sc_norm.str.len() >= 2)
        all_curr_season = all_data[_curr_mask]
        curr_year_codes = sorted(
            {
                str(x).strip().upper()
                for x in all_curr_season["season_code"].dropna().unique()
                if str(x).strip() and len(str(x).strip()) >= 2
            },
            key=_season_code_sort_key,
        )
        category_stockout_by_season: dict[str, list] = {
            "all": _category_stats_rows(all_curr_season),
        }
        for code in curr_year_codes:
            category_stockout_by_season[code] = _category_stats_rows(all_data.loc[_sc_norm == code])

        sc_series = all_data["season_code"].astype(str).str.strip()
        valid_season_mask = all_data["season_code"].notna() & (sc_series.str.len() >= 2)
        sub_season = all_data.loc[valid_season_mask].copy()
        if sub_season.empty:
            season_stockout_stats = []
        else:
            sub_season["_sc"] = (
                all_data.loc[valid_season_mask, "season_code"].astype(str).str.strip()
            )
            sg = sub_season.groupby("_sc", as_index=False).agg(
                total=("stock", "count"),
                stockout=("stock", lambda s: int((s == 0).sum())),
            )
            sg["rate"] = (sg["stockout"] / sg["total"] * 100).where(sg["total"] > 0, 0).round(1)
            sg["letter"] = sg["_sc"].str[0].str.upper()
            sg = sg.sort_values("_sc")
            season_stockout_stats = [
                {
                    "code": str(r["_sc"]),
                    "total": int(r["total"]),
                    "stockout": int(r["stockout"]),
                    "rate": float(r["rate"]),
                    "letter": str(r["letter"]),
                }
                for _, r in sg.iterrows()
            ]

        season_groups = defaultdict(list)
        for srow in season_stockout_stats:
            season_groups[srow["letter"]].append(srow)

        season_group_stats = []
        for letter in sorted(season_groups.keys()):
            seasons = season_groups[letter]
            group_total = sum(se["total"] for se in seasons)
            group_stockout = sum(se["stockout"] for se in seasons)
            group_rate = round((group_stockout / group_total * 100), 1) if group_total > 0 else 0.0
            season_group_stats.append({
                "letter": letter,
                "total": group_total,
                "stockout": group_stockout,
                "rate": group_rate,
                "seasons": sorted(seasons, key=lambda x: x["code"]),
                "is_current": letter == current_season_letter,
                "is_prev": letter == prev_season_letter,
            })

        urgent_categories = ["(전체)"] + sorted(all_data["category_code"].dropna().unique().tolist())

        omni_summary = None
        omni_table = []
        try:
            omni_df = pd.read_sql_query(
                """
                SELECT style_code, sku_code, blocked_qty, top_store
                FROM omni_blocked
                WHERE snapshot_date = ?
                """,
                conn,
                params=(latest_date,),
            )
            if not omni_df.empty:
                omni_join = omni_df.merge(
                    all_data[["sku", "product_code", "name", "sales_qty", "stock"]],
                    left_on="sku_code",
                    right_on="sku",
                    how="left",
                )
                style_count = int(omni_join["style_code"].nunique())
                blocked_total = int(omni_join["blocked_qty"].sum())
                store_count = int(
                    omni_join["top_store"].fillna("").replace("", pd.NA).dropna().nunique()
                )
                omni_summary = {
                    "style_count": style_count,
                    "blocked_total": blocked_total,
                    "store_count": store_count,
                }
                omni_view = omni_join.copy()
                omni_view["sales_qty"] = omni_view["sales_qty"].fillna(0).astype(int)
                omni_view["stock"] = omni_view["stock"].fillna(0).astype(int)
                _sn = all_data[["product_code", "name"]].copy()
                _sn["pc"] = _sn["product_code"].astype(str).str.strip()
                _sn["nm"] = _sn["name"].fillna("").astype(str).str.strip()
                _sn = _sn[(_sn["pc"] != "") & (_sn["nm"] != "")]
                style_name_lookup = _sn.drop_duplicates(subset=["pc"], keep="first").set_index("pc")["nm"].to_dict()
                omni_view["_sk"] = omni_view["style_code"].astype(str).str.strip()
                omni_view["style_name"] = omni_view["name"].fillna("").astype(str).str.strip()
                _miss = omni_view["style_name"] == ""
                omni_view.loc[_miss, "style_name"] = omni_view.loc[_miss, "_sk"].map(
                    lambda k: style_name_lookup.get(k, "") if k else ""
                )
                omni_view = omni_view.drop(columns=["_sk"])
                omni_view = omni_view.sort_values("blocked_qty", ascending=False)
                omni_table = [
                    {
                        "style_code": str(r["style_code"]),
                        "style_name": str(r.get("style_name") or "").strip(),
                        "sku_code": str(r["sku_code"]),
                        "blocked_qty": int(r["blocked_qty"]),
                        "sales_qty": int(r["sales_qty"]),
                        "stock": int(r["stock"]),
                        "top_store": (r.get("top_store") or ""),
                    }
                    for r in omni_view.to_dict(orient="records")
                ]
        except Exception as ex:
            print(f"[WARN] 옴니판매불가 데이터 로딩 실패: {ex}")
            omni_summary = None
            omni_table = []

        categories = [
            "(전체)", "긴급필업", "재고없음", "필업필요", "체크필요", "저재고", "필업검토", "정상",
        ]
        season_codes = ["(전체)"] + sorted(all_data["season_code"].dropna().unique().tolist())
        kpi_all = {
            "total_items": total_items_all,
            "total_stock": total_stock_all,
            "oos": oos_all,
            "low": low_all,
            "has_channel": has_channel_all,
            "total_channel_stock": total_channel_stock_all,
            "has_warehouse": has_warehouse_all,
            "total_warehouse_stock": total_warehouse_stock_all,
            "stockout_count": stockout_count_curr_year,
            "stockout_rate": stockout_rate_curr_year,
            "stockout_denominator": total_items_curr_year,
            "distribution_items": distribution_items_all,
            "distribution_total_qty": distribution_total_qty_all,
        }
        out = {
            "latest_date": latest_date,
            "latest": latest.copy(),
            "all_data": all_data,
            "kpi_all": kpi_all,
            "category_stockout_by_season": category_stockout_by_season,
            "curr_year_codes": curr_year_codes,
            "current_season_letter": current_season_letter,
            "prev_season_letter": prev_season_letter,
            "season_group_stats": season_group_stats,
            "season_codes": season_codes,
            "urgent_categories": urgent_categories,
            "categories": categories,
            "omni_summary": omni_summary,
            "omni_table": omni_table,
        }
        cache.set("dashboard_base", out, timeout=300)
        return out
    finally:
        conn.close()


def _dashboard_impl():
    # [최적화] _load_base_data 캐시 + 요청별 suggested_order_qty·긴급·필터·아이템·차트
    """대시보드 로직 구현"""
    base = _load_base_data()
    if base.get("latest_date") is None:
        return render_template("empty.html", title=APP_TITLE)

    latest_date = base["latest_date"]
    latest = base["latest"]
    category = (request.args.get("category") or "(전체)").strip()
    q = (request.args.get("q") or "").strip()
    low_only = (request.args.get("low_only") or "0").strip() == "1"
    warehouse_only = (request.args.get("warehouse_only") or "0").strip() == "1"
    channel_only = (request.args.get("channel_only") or "0").strip() == "1"
    distribution_only = (request.args.get("distribution_only") or "0").strip() == "1"
    warehouse_center = (request.args.get("warehouse_center") or "전체").strip()
    season_codes_selected = request.args.getlist("season_code")
    urgent_category = (request.args.get("urgent_category") or "(전체)").strip()
    target_cover_days = int((request.args.get("target_cover_days") or "14").strip() or 14)
    sku_pick: Optional[str] = (request.args.get("sku") or "").strip() or None
    page = max(1, int(request.args.get("page", 1) or 1))

    working = base["all_data"]
    working = working.copy()
    working["suggested_order_qty"] = (
        (working["daily_sales_7d"] * target_cover_days) - working["total_available"]
    ).clip(lower=0).astype(int)

    high_risk_all = working[
        (working["daily_sales_7d"] > 0)
        & (
            (working["status"].isin(["긴급필업", "필업필요", "체크필요"]))
            | (working["days_until_out"] < 14)
        )
    ].copy()
    if urgent_category != "(전체)":
        high_risk_all = high_risk_all[high_risk_all["category_code"] == urgent_category]
    high_risk_summary = (
        high_risk_all.sort_values("daily_sales_7d", ascending=False).head(30).to_dict(orient="records")
        if not high_risk_all.empty
        else []
    )

    kpi_all = base["kpi_all"]
    category_stockout_by_season = base["category_stockout_by_season"]
    curr_year_codes = base["curr_year_codes"]
    current_season_letter = base["current_season_letter"]
    prev_season_letter = base["prev_season_letter"]
    season_group_stats = base["season_group_stats"]
    categories = base["categories"]
    season_codes = base["season_codes"]
    urgent_categories = base["urgent_categories"]
    omni_summary = base["omni_summary"]
    omni_table = base["omni_table"]

    view = working
    if q:
        or_groups = [g.strip() for g in q.split(",") if g.strip()]
        vq = view.copy()
        vq["_sku_l"] = vq["sku"].astype(str).str.lower()
        vq["_nm_l"] = vq["name"].fillna("").astype(str).str.lower()
        final_mask = pd.Series([False] * len(vq), index=vq.index)
        for group in or_groups:
            and_terms = group.lower().split()
            group_mask = pd.Series([True] * len(vq), index=vq.index)
            for term in and_terms:
                term_mask = vq["_sku_l"].str.contains(term, na=False, regex=False) | vq[
                    "_nm_l"
                ].str.contains(term, na=False, regex=False)
                group_mask = group_mask & term_mask
            final_mask = final_mask | group_mask
        view = vq.loc[final_mask].drop(columns=["_sku_l", "_nm_l"], errors="ignore")

    if category != "(전체)":
        view = view[view["status"] == category]

    if season_codes_selected and len(season_codes_selected) > 0:
        view = view[view["season_code"].isin(season_codes_selected)]

    if low_only:
        view = view[view["status"].isin(["긴급필업", "재고없음", "필업필요", "체크필요", "저재고", "필업검토"])]

    if warehouse_only:
        view = view[view["warehouse_stock"] > 0]

    if channel_only:
        view = view[view["channel_stock"] > 0]

    if distribution_only:
        view = view[view["distribution_note"].fillna("").astype(str).str.strip() != ""]

    if warehouse_center == "센터1":
        view = view[view["warehouse1_stock"] > 0]
    elif warehouse_center == "센터2":
        view = view[view["warehouse2_stock"] > 0]

    view["avg_daily_usage_est"] = 0.0

    filtered_items = int(view["sku"].nunique())
    filtered_stockout_count = int((view["stock"] == 0).sum())
    filtered_stockout_rate = round((filtered_stockout_count / filtered_items * 100), 1) if filtered_items > 0 else 0.0
    filtered_warehouse_available_count = int((view["warehouse_stock"] > 0).sum())
    filtered_warehouse_available_pct = round(
        (filtered_warehouse_available_count / filtered_items * 100), 1
    ) if filtered_items > 0 else 0.0

    table_columns = [
        "status", "product_code", "sku", "name", "category", "stock", "channel_stock",
        "warehouse1_stock", "warehouse2_stock", "warehouse_stock",
        "daily_sales_7d", "days_until_out", "suggested_order_qty", "distribution_note",
        "min_stock", "reorder_point", "avg_daily_usage_est",
        "lead_time_days", "safety_stock",
    ]

    PAGE_SIZE = 300
    sorted_view = view[table_columns].sort_values("daily_sales_7d", ascending=False)
    total_count = len(sorted_view)
    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(1, min(page, total_pages))
    table = sorted_view.iloc[(page - 1) * PAGE_SIZE : page * PAGE_SIZE].to_dict(orient="records")

    def _dash_pagination_url(target_page: int) -> str:
        pairs: list[tuple[str, str]] = []
        for key in request.args:
            if key == "page":
                continue
            for val in request.args.getlist(key):
                pairs.append((key, val))
        pairs.append(("page", str(target_page)))
        return f"{url_for('dashboard')}?{urlencode(pairs, doseq=True)}"

    pagination_prev_url = _dash_pagination_url(page - 1) if page > 1 else ""
    pagination_next_url = _dash_pagination_url(page + 1) if page < total_pages else ""

    item_tab_req = (request.args.get("item_tab") or "all").strip().upper()
    if item_tab_req != "ALL" and item_tab_req not in {str(c).upper() for c in curr_year_codes}:
        item_tab_req = "ALL"
    item_tab_selected = item_tab_req.lower()

    conn = get_conn()
    try:
        item_summary, item_prev_date, item_has_prev = _build_item_inventory_summary(
            conn,
            str(latest_date),
            latest,
            current_season_letter,
            item_tab_req,
            curr_year_codes,
        )

        sku_list = sorted(latest["sku"].astype(str).unique().tolist())
        sku_pick = sku_pick or (sku_list[0] if sku_list else None)
        chart_sku_line_html = None
        chart_sku_delta_html = None

        if sku_pick:
            hist = load_history(conn, sku_pick)
            if len(hist) >= 2:
                h = compute_daily_change(hist)
                h["snapshot_date"] = pd.to_datetime(h["snapshot_date"])

                fig_line = px.line(h, x="snapshot_date", y="stock", markers=True, title=f"SKU {sku_pick} 재고 변동")
                fig_line.update_layout(height=300)

                fig_delta = px.bar(h.dropna(subset=["delta"]), x="snapshot_date", y="delta", title="일별 재고 증감")
                fig_delta.update_layout(height=250)

                chart_sku_line_html = fig_line.to_json()
                chart_sku_delta_html = fig_delta.to_json()
    finally:
        conn.close()

    kpi_filtered = {
        "total_items": filtered_items,
        "stockout_count": filtered_stockout_count,
        "stockout_rate": filtered_stockout_rate,
        "warehouse_available_count": filtered_warehouse_available_count,
        "warehouse_available_pct": filtered_warehouse_available_pct,
    }

    return render_template(
        "dashboard.html",
        title=APP_TITLE,
        latest_date=latest_date,
        kpi=kpi_all,
        kpi_filtered=kpi_filtered,
        category_stockout_by_season=category_stockout_by_season,
        curr_year_season_codes=curr_year_codes,
        current_season_letter=current_season_letter,
        prev_season_letter=prev_season_letter,
        season_group_stats=season_group_stats,
        categories=categories,
        season_codes=season_codes,
        urgent_categories=urgent_categories,
        selected={
            "category": category,
            "q": q,
            "low_only": low_only,
            "warehouse_only": warehouse_only,
            "channel_only": channel_only,
            "distribution_only": distribution_only,
            "warehouse_center": warehouse_center,
            "season_codes": season_codes_selected,
            "urgent_category": urgent_category,
            "target_cover_days": target_cover_days,
            "sku": sku_pick,
            "item_tab": item_tab_selected,
        },
        high_risk_summary=high_risk_summary,
        table=table,
        omni_summary=omni_summary,
        omni_table=omni_table,
        status_badge=_status_badge,
        sku_list=sku_list,
        chart_sku_line_html=chart_sku_line_html,
        chart_sku_delta_html=chart_sku_delta_html,
        item_summary=item_summary,
        item_prev_date=item_prev_date,
        item_has_prev=item_has_prev,
        item_tab_season_codes=curr_year_codes,
        page=page,
        total_pages=total_pages,
        total_count=total_count,
        pagination_prev_url=pagination_prev_url,
        pagination_next_url=pagination_next_url,
    )


@app.route("/download_failed")
@login_required
def download_failed():
    """업로드 실패 목록 다운로드"""
    if 'failed_csv_path' not in session:
        flash("다운로드할 실패 목록이 없습니다.", "warning")
        return redirect(url_for("dashboard"))
    
    csv_path = session['failed_csv_path']
    if not os.path.exists(csv_path):
        flash("실패 목록 파일을 찾을 수 없습니다.", "danger")
        return redirect(url_for("dashboard"))
    
    return send_file(
        csv_path,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'업로드실패목록_{dt.date.today()}.csv'
    )


@app.route("/clear_data", methods=["GET", "POST"])
@login_required
def clear_data():
    """데이터 초기화 (배포 모드에서는 비활성화)"""
    if DEPLOY_MODE:
        abort(404)
    if request.method == "GET":
        # 확인 페이지 표시
        conn = get_conn()
        total_count = pd.read_sql_query("SELECT COUNT(*) as cnt FROM snapshots", conn).iloc[0]['cnt']
        dates_count = pd.read_sql_query("SELECT COUNT(DISTINCT snapshot_date) as cnt FROM snapshots", conn).iloc[0]['cnt']
        return render_template(
            "clear_data.html", 
            title="데이터 초기화",
            total_count=total_count,
            dates_count=dates_count
        )
    
    # POST 요청: 실제 삭제
    confirm = request.form.get("confirm")
    if confirm == "DELETE":
        try:
            conn = get_conn()
            conn.execute("DELETE FROM snapshots")
            conn.commit()
            _invalidate_snapshot_caches()
            flash("✅ 모든 데이터가 삭제되었습니다.", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            flash(f"❌ 데이터 삭제 실패: {e}", "danger")
            return redirect(url_for("clear_data"))
    else:
        flash("⚠️ 확인 문구가 일치하지 않습니다.", "warning")
        return redirect(url_for("clear_data"))


if __name__ == "__main__":
    import sys
    # Streamlit Cloud 등에서 streamlit run으로 실행될 때는 Flask 서버를 띄우지 않음
    if "streamlit" in sys.modules:
        # Streamlit 전용 앱은 app.py 사용: streamlit run app.py
        pass
    else:
        print("=" * 70)
        print("재고 대시보드 V4 서버 시작!")
        print("=" * 70)
        print("접속 주소: http://127.0.0.1:5003")
        print("기본 비밀번호: 1234")
        if DEPLOY_MODE:
            print("모드: 배포 (초기화 비노출, /test 비노출)")
        print("=" * 70)
        print("")
        app.run(host="127.0.0.1", port=5003, debug=not DEPLOY_MODE)
